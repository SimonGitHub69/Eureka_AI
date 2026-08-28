import re
from types import SimpleNamespace
from unittest.mock import MagicMock, patch

from django.test import SimpleTestCase, override_settings

from apps.core import ai_assistant


class AiAssistantPromptGuardTests(SimpleTestCase):
    def test_detects_article_category_request(self):
        self.assertTrue(
            ai_assistant._is_article_category_request(
                "mostrami gli articoli del settore abbigliamento"
            )
        )

    def test_does_not_mark_plain_article_search_as_category_request(self):
        self.assertFalse(
            ai_assistant._is_article_category_request(
                "cerca articoli con vite nella descrizione"
            )
        )

    def test_detects_explicit_article_text_search(self):
        self.assertTrue(
            ai_assistant._is_explicit_article_text_search(
                "cerca articoli con abbigliamento nella descrizione"
            )
        )

    def test_detects_synonym_description_search(self):
        self.assertTrue(
            ai_assistant._is_synonym_description_search(
                "articoli con descrizione sinonimi di calzature"
            )
        )

    def test_build_user_prompt_adds_synonym_hint_for_description_synonyms(self):
        prompt = "articoli con descrizione sinonimi di calzature"
        guarded = ai_assistant._build_ai_user_prompt(prompt)
        self.assertIn("sinonimi in descrizione", guarded)
        self.assertIn('articoli."Descrizione"', guarded)
        self.assertIn("ILIKE", guarded)
        self.assertIn("calzature", guarded)
        self.assertIn("scarpe", guarded)

    def test_detects_inactive_article_status_request(self):
        self.assertTrue(
            ai_assistant._is_article_inactive_status_request(
                "articoli disattivati con descrizione sinonimi di calzature"
            )
        )

    def test_detects_active_article_status_request(self):
        self.assertTrue(
            ai_assistant._is_article_active_status_request(
                "mostrami gli articoli attivi del gruppo GR10"
            )
        )

    def test_build_user_prompt_adds_disattivati_sql_hint(self):
        prompt = "articoli disattivati con descrizione sinonimi di calzature"
        guarded = ai_assistant._build_ai_user_prompt(prompt)
        self.assertIn('"FlDisattivato" IS TRUE', guarded)
        self.assertIn("NON usare = true", guarded)
        self.assertIn("sinonimi in descrizione", guarded)

    def test_build_user_prompt_adds_attivi_sql_hint(self):
        prompt = "articoli attivi con calzature nella descrizione"
        guarded = ai_assistant._build_ai_user_prompt(prompt)
        self.assertIn('"FlDisattivato" IS NOT TRUE', guarded)
        self.assertIn("NON usare = true", guarded)

    def test_build_user_prompt_adds_structured_guard_for_category_requests(self):
        guarded = ai_assistant._build_ai_user_prompt(
            "mostrami gli articoli della categoria abbigliamento"
        )
        self.assertIn('articoli."CatOmogenea"', guarded)
        self.assertIn('articoli."CodGruppo"', guarded)
        self.assertIn('{"sql": null, "spiegazione": "..."}', guarded)
        self.assertIn('Non usare articoli."Descrizione"', guarded)

    def test_build_user_prompt_keeps_explicit_text_search_unchanged(self):
        prompt = "cerca articoli con abbigliamento nella descrizione"
        self.assertEqual(ai_assistant._build_ai_user_prompt(prompt), prompt)

    @override_settings(AI_BACKEND="openai")
    @patch("apps.core.ai_assistant._execute_query", return_value=([], False))
    @patch("apps.core.ai_assistant._get_model", return_value="test-model")
    @patch("apps.core.ai_assistant._get_client")
    def test_ask_ai_sends_guarded_prompt_for_article_category_requests(
        self,
        get_client_mock,
        _get_model_mock,
        _execute_query_mock,
    ):
        fake_client = SimpleNamespace()
        fake_client.chat = SimpleNamespace()
        fake_client.chat.completions = SimpleNamespace()
        fake_client.chat.completions.create = lambda **kwargs: SimpleNamespace(
            choices=[
                SimpleNamespace(
                    message=SimpleNamespace(
                        content='{"sql":"SELECT \\"Codice\\", \\"CatOmogenea\\" FROM articoli LIMIT 1","spiegazione":"ok"}'
                    )
                )
            ]
        )
        get_client_mock.return_value = fake_client

        ai_assistant.ask_ai("mostrami gli articoli del settore abbigliamento")

        create_call = get_client_mock.return_value.chat.completions.create
        self.assertIsNotNone(create_call)

    @override_settings(AI_BACKEND="openai")
    @patch("apps.core.ai_assistant._execute_query", return_value=([], False))
    @patch("apps.core.ai_assistant._get_model", return_value="test-model")
    @patch("apps.core.ai_assistant._get_client")
    def test_ask_ai_message_contains_structured_guard(
        self,
        get_client_mock,
        _get_model_mock,
        _execute_query_mock,
    ):
        recorded = {}

        def fake_create(**kwargs):
            recorded["messages"] = kwargs["messages"]
            return SimpleNamespace(
                choices=[
                    SimpleNamespace(
                        message=SimpleNamespace(
                            content='{"sql":"SELECT \\"Codice\\", \\"CatOmogenea\\" FROM articoli LIMIT 1","spiegazione":"ok"}'
                        )
                    )
                ]
            )

        fake_client = SimpleNamespace()
        fake_client.chat = SimpleNamespace()
        fake_client.chat.completions = SimpleNamespace(create=fake_create)
        get_client_mock.return_value = fake_client

        ai_assistant.ask_ai("mostrami gli articoli del settore abbigliamento")

        user_message = recorded["messages"][1]["content"]
        self.assertIn('articoli."CatOmogenea"', user_message)
        self.assertIn('articoli."CodGruppo"', user_message)
        self.assertIn("Non usare articoli.", user_message)

    @override_settings(AI_BACKEND="openai")
    @patch("apps.core.ai_assistant._try_fast_path_articoli_sql", return_value=None)
    @patch("apps.core.ai_assistant._execute_query", return_value=([], False))
    @patch("apps.core.ai_assistant._get_model", return_value="test-model")
    @patch("apps.core.ai_assistant._get_client")
    def test_ask_ai_keeps_explicit_description_search_prompt(
        self,
        get_client_mock,
        _get_model_mock,
        _execute_query_mock,
        _fast_path_mock,
    ):
        recorded = {}

        def fake_create(**kwargs):
            recorded["messages"] = kwargs["messages"]
            return SimpleNamespace(
                choices=[
                    SimpleNamespace(
                        message=SimpleNamespace(
                            content='{"sql":"SELECT \\"Codice\\", \\"Descrizione\\" FROM articoli WHERE \\"Descrizione\\" ILIKE \'%abbigliamento%\' LIMIT 1","spiegazione":"ok"}'
                        )
                    )
                ]
            )

        fake_client = SimpleNamespace()
        fake_client.chat = SimpleNamespace()
        fake_client.chat.completions = SimpleNamespace(create=fake_create)
        get_client_mock.return_value = fake_client

        prompt = "cerca articoli con abbigliamento nella descrizione"
        ai_assistant.ask_ai(prompt)

        self.assertEqual(recorded["messages"][1]["content"], prompt)

    @override_settings(AI_BACKEND="openai")
    @patch("apps.core.ai_assistant._try_fast_path_articoli_sql", return_value=None)
    @patch("apps.core.ai_assistant._execute_query", return_value=([], False))
    @patch("apps.core.ai_assistant._get_model", return_value="test-model")
    @patch("apps.core.ai_assistant._get_client")
    def test_ask_ai_sends_synonym_hint_for_description_synonyms(
        self,
        get_client_mock,
        _get_model_mock,
        _execute_query_mock,
        _fast_path_mock,
    ):
        recorded = {}

        def fake_create(**kwargs):
            recorded["messages"] = kwargs["messages"]
            return SimpleNamespace(
                choices=[
                    SimpleNamespace(
                        message=SimpleNamespace(
                            content=(
                                '{"sql":"SELECT \\"Codice\\", \\"Descrizione\\" FROM articoli WHERE '
                                '(\\"Descrizione\\" ILIKE \'%calzature%\' OR \\"Descrizione\\" ILIKE \'%scarpe%\') '
                                'LIMIT 5","spiegazione":"ok"}'
                            )
                        )
                    )
                ]
            )

        fake_client = SimpleNamespace()
        fake_client.chat = SimpleNamespace()
        fake_client.chat.completions = SimpleNamespace(create=fake_create)
        get_client_mock.return_value = fake_client

        prompt = "articoli con descrizione sinonimi di calzature"
        result = ai_assistant.ask_ai(prompt, limit=5)

        user_message = recorded["messages"][1]["content"]
        self.assertIn("sinonimi in descrizione", user_message)
        self.assertTrue(result["ok"])
        self.assertIn("ILIKE", result["sql"])


class AiAssistantFastPathTests(SimpleTestCase):
    def test_extracts_search_term_from_synonym_prompt(self):
        term = ai_assistant._extract_article_search_term(
            "articoli disattivati con descrizione sinonimi di calzature"
        )
        self.assertEqual(term, "calzature")

    def test_fast_path_generates_disattivati_synonym_sql(self):
        result = ai_assistant._try_fast_path_articoli_sql(
            "articoli disattivati con descrizione sinonimi di calzature"
        )
        self.assertIsNotNone(result)
        sql, spiegazione = result
        self.assertIn('"FlDisattivato" IS TRUE', sql)
        self.assertIn('ILIKE \'%calzature%\'', sql)
        self.assertIn('ILIKE \'%scarpe%\'', sql)
        self.assertIn("disattivati", spiegazione.lower())

    def test_fast_path_caps_synonym_or_clauses(self):
        ai_assistant._ARTICLE_SYNONYM_MAP["testterm"] = [
            f"term{i}" for i in range(20)
        ]
        result = ai_assistant._try_fast_path_articoli_sql(
            "articoli con descrizione sinonimi di testterm"
        )
        self.assertIsNotNone(result)
        sql, _ = result
        self.assertLessEqual(sql.count(" OR "), ai_assistant.MAX_SYNONYM_OR_CLAUSES - 1)

    @patch("apps.core.ai_assistant._get_client")
    @patch("apps.core.ai_assistant._call_llm")
    def test_ask_ai_uses_fast_path_without_llm(
        self,
        call_llm_mock,
        get_client_mock,
    ):
        with patch(
            "apps.core.ai_assistant._execute_query",
            return_value=([{"Codice": "A1", "Descrizione": "Scarpe"}], False),
        ):
            result = ai_assistant.ask_ai(
                "articoli disattivati con descrizione sinonimi di calzature",
                limit=10,
            )
        self.assertTrue(result["ok"])
        self.assertTrue(result["fast_path"])
        self.assertEqual(result["time_spent_llm_request"], 0.0)
        call_llm_mock.assert_not_called()
        get_client_mock.assert_not_called()

    def test_fast_path_still_matches_export_prompt(self):
        result = ai_assistant._try_fast_path_articoli_sql(
            "articoli con descrizione sinonimi di calzature genera un file xlsx"
        )
        self.assertIsNotNone(result)
        sql, _ = result
        self.assertIn('ILIKE \'%calzature%\'', sql)
        self.assertIn('SELECT "Codice", "Descrizione"', sql)


class AiAssistantExportColumnTests(SimpleTestCase):
    PROMPT_EXTRA_FIELDS = (
        "articoli con descrizione sinonimi di calzature "
        "genera xlsx con codice descrizione e prezzo"
    )
    PROMPT_CAMPI_LIST = (
        "esporta articoli calzature con campi Codice, Descrizione, FlDisattivato"
    )
    PROMPT_FORNITORE_EXPORT = (
        "articoli con descrizione sinonimi di calzature "
        "crea un file xlsx con codice, descrizione, codice fornitore, "
        "ragione sociale fornitore"
    )

    def test_extract_export_columns_from_campi_list(self):
        from apps.core.ai_export import _extract_export_columns

        columns, unknown, _overrides = _extract_export_columns(self.PROMPT_CAMPI_LIST, "articoli")
        self.assertEqual(columns, ["Codice", "Descrizione", "FlDisattivato"])
        self.assertEqual(unknown, [])

    def test_extract_export_columns_from_xlsx_con_prezzo(self):
        from apps.core.ai_export import _extract_export_columns

        columns, unknown, _overrides = _extract_export_columns(self.PROMPT_EXTRA_FIELDS, "articoli")
        self.assertEqual(columns, ["Codice", "Descrizione", "Listino1"])
        self.assertEqual(unknown, [])

    def test_extract_export_columns_includi_pattern(self):
        from apps.core.ai_export import _extract_export_columns

        prompt = "articoli calzature includi Codice, Descrizione, PrezzoUltCar genera xlsx"
        columns, unknown, _overrides = _extract_export_columns(prompt, "articoli")
        self.assertEqual(columns, ["Codice", "Descrizione", "PrezzoUltCar"])
        self.assertEqual(unknown, [])

    def test_extract_export_columns_ignores_unknown_with_warning(self):
        from apps.core.ai_export import _extract_export_columns, format_export_columns_warning

        columns, unknown, _overrides = _extract_export_columns(
            "esporta articoli con campi Codice, CampoInventato", "articoli"
        )
        self.assertEqual(columns, ["Codice"])
        self.assertEqual(unknown, ["CampoInventato"])
        self.assertIn("CampoInventato", format_export_columns_warning(unknown))

    def test_default_columns_when_export_without_field_list(self):
        from apps.core.ai_export import resolve_export_columns

        prompt = "articoli con descrizione sinonimi di calzature genera un file xlsx"
        columns, unknown, overrides = resolve_export_columns(prompt, "articoli", for_export=True)
        self.assertEqual(columns, ["Codice", "Descrizione"])
        self.assertEqual(unknown, [])
        self.assertEqual(overrides, {})

    def test_resolve_prezzo_alias_to_listino1(self):
        from apps.core.ai_export import resolve_column_name

        self.assertEqual(resolve_column_name("prezzo", "articoli"), "Listino1")
        self.assertEqual(resolve_column_name("FlDisattivato", "articoli"), "FlDisattivato")

    def test_fast_path_sql_includes_extra_columns(self):
        result = ai_assistant._try_fast_path_articoli_sql(self.PROMPT_EXTRA_FIELDS)
        self.assertIsNotNone(result)
        sql, _ = result
        self.assertIn('"Codice"', sql)
        self.assertIn('"Descrizione"', sql)
        self.assertIn('"Listino1"', sql)

    def test_resolve_fornitore_column_aliases(self):
        from apps.core.ai_export import resolve_column_name

        self.assertEqual(resolve_column_name("codice fornitore", "articoli"), "CodFornitore")
        self.assertEqual(
            resolve_column_name("ragione sociale fornitore", "articoli"),
            "RagioneSocialeFornitore",
        )

    def test_resolve_column_names_case_insensitive(self):
        from apps.core.ai_export import resolve_column_name

        self.assertEqual(resolve_column_name("CODICE", "articoli"), "Codice")
        self.assertEqual(resolve_column_name("Codice Fornitore", "articoli"), "CodFornitore")
        self.assertEqual(
            resolve_column_name("RAGIONE SOCIALE FORNITORE", "articoli"),
            "RagioneSocialeFornitore",
        )
        self.assertEqual(resolve_column_name("PREZZO", "articoli"), "Listino1")
        self.assertEqual(resolve_column_name("FlDisattivato", "articoli"), "FlDisattivato")
        self.assertEqual(resolve_column_name("fldisattivato", "articoli"), "FlDisattivato")

    def test_extract_export_columns_uppercase_space_separated(self):
        from apps.core.ai_export import _extract_export_columns

        prompt = (
            "articoli con descrizione sinonimi di calzature "
            "genera xlsx con CODICE DESCRIZIONE PREZZO"
        )
        columns, unknown, _overrides = _extract_export_columns(prompt, "articoli")
        self.assertEqual(columns, ["Codice", "Descrizione", "Listino1"])
        self.assertEqual(unknown, [])

    def test_extract_export_columns_mixed_case_fornitore_aliases(self):
        from apps.core.ai_export import _extract_export_columns

        prompt = (
            "esporta articoli con campi: CODICE, Codice Fornitore, "
            "RAGIONE SOCIALE FORNITORE"
        )
        columns, unknown, _overrides = _extract_export_columns(prompt, "articoli")
        self.assertEqual(
            columns,
            ["Codice", "CodFornitore", "RagioneSocialeFornitore"],
        )
        self.assertEqual(unknown, [])

    def test_wants_xlsx_export_case_insensitive(self):
        self.assertTrue(ai_assistant._wants_xlsx_export("ESPORTA articoli in EXCEL"))
        self.assertTrue(ai_assistant._wants_xlsx_export("Genera File XLSX con codice"))

    def test_ensure_case_insensitive_text_search_converts_like(self):
        sql = (
            'SELECT "Codice" FROM articoli WHERE descrizione LIKE \'%calzature%\' '
            "AND descrizione NOT LIKE '%obsolete%'"
        )
        fixed = ai_assistant._ensure_case_insensitive_text_search(
            ai_assistant._fix_column_quoting(sql)
        )
        self.assertNotIn(" LIKE ", fixed.upper().replace("ILIKE", "X"))
        self.assertIn("ILIKE '%calzature%'", fixed)
        self.assertIn("NOT ILIKE '%obsolete%'", fixed)
        self.assertIn('"Descrizione"', fixed)

    def test_export_sql_uses_ilike_after_normalization(self):
        raw_sql = (
            'SELECT "Codice", "Descrizione" FROM articoli WHERE '
            '"Descrizione" LIKE \'%calzature%\''
        )
        with patch(
            "apps.core.ai_assistant._execute_query",
            return_value=([{"Codice": "A1", "Descrizione": "Scarpa"}], False),
        ):
            with patch(
                "apps.core.ai_assistant._try_fast_path_sql",
                return_value=(raw_sql, "test"),
            ):
                result = ai_assistant.ask_ai(
                    "articoli calzature genera file xlsx con codice descrizione",
                    limit=10,
                )
        self.assertTrue(result["ok"])
        self.assertIn("ILIKE", result["sql"])
        self.assertNotRegex(result["sql"], r"(?<!I)\bLIKE\b", re.IGNORECASE)

    def test_extract_export_columns_fornitore_fields(self):
        from apps.core.ai_export import _extract_export_columns

        columns, unknown, _overrides = _extract_export_columns(self.PROMPT_FORNITORE_EXPORT, "articoli")
        self.assertEqual(
            columns,
            ["Codice", "Descrizione", "CodFornitore", "RagioneSocialeFornitore"],
        )
        self.assertEqual(unknown, [])

    def test_column_display_labels_fornitore(self):
        from apps.core.ai_export import column_display_labels

        labels = column_display_labels(
            "articoli",
            ["Codice", "Descrizione", "CodFornitore", "RagioneSocialeFornitore"],
        )
        self.assertEqual(labels[2], "Codice fornitore")
        self.assertEqual(labels[3], "Ragione sociale fornitore")

    def test_extract_export_columns_with_header_aliases(self):
        from apps.core.ai_export import (
            _extract_export_columns,
            _parse_export_field_token,
            column_display_labels,
        )

        self.assertEqual(
            _parse_export_field_token("codice come Cod Art"),
            ("codice", "Cod Art"),
        )
        self.assertEqual(
            _parse_export_field_token("descrizione as Descrizione breve"),
            ("descrizione", "Descrizione breve"),
        )
        self.assertEqual(
            _parse_export_field_token("prezzo alias Prezzo vendita"),
            ("prezzo", "Prezzo vendita"),
        )

        prompt = (
            "articoli calzature genera xlsx con "
            "codice come Cod Art, descrizione as Descrizione breve, "
            "ragione sociale fornitore alias Fornitore"
        )
        columns, unknown, overrides = _extract_export_columns(prompt, "articoli")
        self.assertEqual(columns, ["Codice", "Descrizione", "RagioneSocialeFornitore"])
        self.assertEqual(unknown, [])
        self.assertEqual(
            overrides,
            {
                "Codice": "Cod Art",
                "Descrizione": "Descrizione breve",
                "RagioneSocialeFornitore": "Fornitore",
            },
        )
        labels = column_display_labels("articoli", columns, overrides)
        self.assertEqual(labels, ["Cod Art", "Descrizione breve", "Fornitore"])

    def test_extract_export_columns_header_aliases_without_commas(self):
        from apps.core.ai_export import _extract_export_columns

        prompt = (
            "articoli calzature genera xlsx con "
            "codice come Cod Art descrizione as Desc breve prezzo alias Listino"
        )
        columns, unknown, overrides = _extract_export_columns(prompt, "articoli")
        self.assertEqual(columns, ["Codice", "Descrizione", "Listino1"])
        self.assertEqual(unknown, [])
        self.assertEqual(
            overrides,
            {
                "Codice": "Cod Art",
                "Descrizione": "Desc breve",
                "Listino1": "Listino",
            },
        )

    def test_fast_path_sql_includes_fornitore_join(self):
        result = ai_assistant._try_fast_path_articoli_sql(self.PROMPT_FORNITORE_EXPORT)
        self.assertIsNotNone(result)
        sql, _ = result
        self.assertIn('LEFT JOIN fornitori', sql)
        self.assertIn('UPPER(fornitori."Codice") = UPPER(articoli."CodFornitore")', sql)
        self.assertIn('articoli."Codice"', sql)
        self.assertIn('articoli."CodFornitore"', sql)
        self.assertIn('"RagioneSocialeFornitore"', sql)
        self.assertIn('articoli."Descrizione"', sql)
        self.assertNotIn('TRIM(BOTH,', sql)

    def test_extract_export_columns_extended_articoli_fields(self):
        from apps.core.ai_export import _extract_export_columns

        prompt = (
            "articoli con descrizione sinonimi di calzature crea un file xlsx con "
            "codice, descrizione, codice fornitore, ragione sociale fornitore, "
            "unitamisura, peso netto, peso lordo, categoria, descrizione_categoria"
        )
        columns, unknown, _overrides = _extract_export_columns(prompt, "articoli")
        self.assertEqual(
            columns,
            [
                "Codice",
                "Descrizione",
                "CodFornitore",
                "RagioneSocialeFornitore",
                "UnitaMisura",
                "PesoNetto",
                "PesoLordo_Manodopera",
                "CatOmogenea",
                "DescrizioneCategoria",
            ],
        )
        self.assertEqual(unknown, [])

    def test_extract_export_columns_csv_with_quoted_header_aliases(self):
        from apps.core.ai_export import _extract_export_columns, column_display_labels

        prompt = (
            "articoli con descrizione sinonimi di calzature crea un file csv con "
            "codice, descrizione, codice fornitore, ragione sociale fornitore, "
            "unitamisura, peso netto, peso lordo alias 'peso_lordo', categoria, "
            "descrizione_categoria, codicealternativo1 as 'Cod.Forn.Art.'"
        )
        columns, unknown, overrides = _extract_export_columns(prompt, "articoli")
        self.assertEqual(
            columns,
            [
                "Codice",
                "Descrizione",
                "CodFornitore",
                "RagioneSocialeFornitore",
                "UnitaMisura",
                "PesoNetto",
                "PesoLordo_Manodopera",
                "CatOmogenea",
                "DescrizioneCategoria",
                "CodiceAlternativo1",
            ],
        )
        self.assertEqual(unknown, [])
        self.assertEqual(overrides["PesoLordo_Manodopera"], "peso_lordo")
        self.assertEqual(overrides["CodiceAlternativo1"], "Cod.Forn.Art.")
        labels = column_display_labels("articoli", columns, overrides)
        self.assertEqual(labels[-1], "Cod.Forn.Art.")
        self.assertIn("peso_lordo", labels)

    def test_fast_path_sql_includes_codice_alternativo_for_csv_export(self):
        prompt = (
            "articoli con descrizione sinonimi di calzature crea un file csv con "
            "codice, descrizione, codicealternativo1 as 'Cod.Forn.Art.'"
        )
        result = ai_assistant._try_fast_path_articoli_sql(prompt)
        self.assertIsNotNone(result)
        sql, _ = result
        self.assertIn('"CodiceAlternativo1"', sql)
        self.assertIn('ILIKE \'%calzature%\'', sql)

    def test_wants_csv_export_and_format(self):
        prompt = (
            "articoli con descrizione sinonimi di calzature crea un file csv con codice"
        )
        self.assertTrue(ai_assistant._wants_xlsx_export(prompt))
        self.assertEqual(ai_assistant._resolve_export_format(prompt), "csv")
        self.assertEqual(
            ai_assistant._resolve_export_format(
                "articoli calzature genera un file xlsx"
            ),
            "xlsx",
        )

    def test_extract_export_columns_agenti(self):
        from apps.core.ai_export import _extract_export_columns, column_display_labels

        prompt = (
            "crea un file dalla tabella Agenti xlsx con codice agente, ragione sociale"
        )
        columns, unknown, overrides = _extract_export_columns(prompt, "agenti")
        self.assertEqual(columns, ["Codice", "RagioneSociale"])
        self.assertEqual(unknown, [])
        labels = column_display_labels("agenti", columns, overrides)
        self.assertEqual(labels, ["Codice agente", "Ragione sociale"])

    def test_fast_path_agenti_export_sql(self):
        prompt = (
            "crea un file dalla tabella Agenti xlsx con codice agente, ragione sociale"
        )
        result = ai_assistant._try_fast_path_table_export_sql(prompt)
        self.assertIsNotNone(result)
        sql, spiegazione = result
        self.assertIn('SELECT "Codice", "RagioneSociale" FROM agenti', sql)
        self.assertIn("Export tabella agenti", spiegazione)

    def test_is_safe_sql_rejects_into_outfile(self):
        sql = (
            'SELECT "Codice" FROM agenti INTO OUTFILE '
            "'/tmp/agenti.xlsx' FIELDS TERMINATED BY ','"
        )
        self.assertFalse(ai_assistant._is_safe_sql(sql))

    @patch("apps.core.ai_assistant._get_client")
    @patch("apps.core.ai_assistant._call_llm")
    def test_ask_ai_agenti_export_uses_fast_path(
        self,
        call_llm_mock,
        get_client_mock,
    ):
        import tempfile
        from io import BytesIO
        from pathlib import Path

        from openpyxl import load_workbook

        prompt = (
            "crea un file dalla tabella Agenti xlsx con codice agente, ragione sociale"
        )
        rows = [
            {"Codice": "AG01", "RagioneSociale": "Rossi Mario"},
            {"Codice": "AG02", "RagioneSociale": "Bianchi Luca"},
        ]
        with tempfile.TemporaryDirectory() as tmp:
            with override_settings(AI_EXPORT_DIR=tmp):
                with patch(
                    "apps.core.ai_assistant._execute_query",
                    return_value=(rows, False),
                ):
                    result = ai_assistant.ask_ai(prompt, limit=10)

            self.assertTrue(result["ok"])
            self.assertTrue(result["fast_path"])
            self.assertTrue(result["export_requested"])
            self.assertIsNotNone(result["export_token"])
            self.assertEqual(
                result["export_headers"],
                ["Codice agente", "Ragione sociale"],
            )
            call_llm_mock.assert_not_called()

            token = result["export_token"]
            xlsx_path = Path(tmp) / f"{token}.xlsx"
            self.assertTrue(xlsx_path.is_file())
            wb = load_workbook(BytesIO(xlsx_path.read_bytes()))
            ws = wb.active
            self.assertEqual(
                [cell.value for cell in ws[1]],
                ["Codice agente", "Ragione sociale"],
            )

    def test_extract_export_columns_pdc(self):
        from apps.core.ai_export import (
            _extract_export_columns,
            detect_export_table_from_prompt,
        )

        prompt = (
            "crea un file dalla tabella PDC in formato xlsx con codice, descrizione"
        )
        self.assertEqual(detect_export_table_from_prompt(prompt), "pdc")
        columns, unknown, _ = _extract_export_columns(prompt, "pdc")
        self.assertEqual(columns, ["Codice", "Descrizione"])
        self.assertEqual(unknown, [])

    def test_fast_path_causali_contabili_export(self):
        from apps.core.ai_export import detect_export_table_from_prompt, resolve_export_columns

        prompt = "crea file xlsx della lista Causali Contabili con codice, descrizione"
        self.assertEqual(detect_export_table_from_prompt(prompt), "causali_contabili")
        columns, unknown, _ = resolve_export_columns(
            prompt, "causali_contabili", for_export=True
        )
        self.assertEqual(columns, ["Codice", "Descrizione"])
        self.assertEqual(unknown, [])
        result = ai_assistant._try_fast_path_table_export_sql(prompt)
        self.assertIsNotNone(result)
        sql, spiegazione = result
        self.assertIn("FROM causali_contabili", sql)
        self.assertIn('"Codice"', sql)
        self.assertIn('"Descrizione"', sql)
        self.assertIn("causali_contabili", spiegazione)

    def test_primanota_codicepartita_alias(self):
        from apps.core.ai_export import resolve_column_name, resolve_export_columns

        self.assertEqual(
            resolve_column_name("codicepartita", "primanota"), "CodicePartita"
        )
        self.assertEqual(
            resolve_column_name("codice partita", "primanota"), "CodicePartita"
        )
        prompt = (
            "export in xlsx tabella Primanota tipo IVA "
            "numero registrazione, data registrazione, codicepartita, causale, "
            "descrizione_causale_contabile, descrizione, contoavere, avere, "
            "descrizione_avere, codice_iva, importo_iva, "
            "somma(avere+importo_iva) as totaledoc dal 1/6/2026"
        )
        columns, unknown, _ = resolve_export_columns(
            prompt, "primanota", for_export=True
        )
        self.assertEqual(unknown, [])
        self.assertIn("CodicePartita", columns)
        self.assertIn("RagioneSocialePartita", columns)
        self.assertLess(columns.index("DataReg"), columns.index("CodicePartita"))
        self.assertLess(
            columns.index("CodicePartita"),
            columns.index("RagioneSocialePartita"),
        )
        self.assertLess(columns.index("CodicePartita"), columns.index("Causale"))
        result = ai_assistant._try_fast_path_table_export_sql(prompt)
        self.assertIsNotNone(result)
        sql, _ = result
        self.assertIn("LEFT JOIN clienti c_partita", sql)
        self.assertIn("LEFT JOIN fornitori f_partita", sql)
        self.assertIn('AS "RagioneSocialePartita"', sql)

    def test_fast_path_primanota_export_iva_totale_doc(self):
        from apps.core.ai_export import resolve_export_columns

        prompt = (
            "export in xlsx tabella Primanota tipo IVA "
            "numero registrazione, data registrazione, causale, "
            "descrizione_causale_contabile, descrizione, contoavere, avere, "
            "descrizione_avere, codice_iva, importo_iva, "
            "somma(avere+importo_iva) as totaledoc dal 1/6/2026"
        )
        columns, unknown, overrides = resolve_export_columns(
            prompt, "primanota", for_export=True
        )
        self.assertEqual(unknown, [])
        self.assertIn("TotaleDoc", columns)
        self.assertIn("CodiceIva", columns)
        self.assertIn("ImportoIva", columns)
        self.assertIn("DescrizioneCausaleContabile", columns)
        self.assertIn("DescrizioneAvere", columns)
        self.assertEqual(overrides.get("TotaleDoc"), "totaledoc")
        result = ai_assistant._try_fast_path_table_export_sql(prompt)
        self.assertIsNotNone(result)
        sql, _ = result
        self.assertIn('p."Tipo" = 2', sql)
        self.assertIn(
            '(COALESCE(pd."Avere_Imponibile", 0) + COALESCE(pd."ImportoIva", 0)) '
            'AS "TotaleDoc"',
            sql,
        )
        self.assertIn("LEFT JOIN causali_contabili cc", sql)
        self.assertIn("LEFT JOIN clienti c_avere", sql)

    def test_fast_path_primanota_export_with_causale_link(self):
        from apps.core.ai_export import resolve_export_columns

        prompt = (
            "export in xlsx tabella Primanota tipo generico "
            "numero registrazione, data registrazione, causale, "
            "descrizione_causale_contabile, descrizione, contodare, dare, "
            "descrizione_dare, contoavere, avere, descrizione_avere "
            "dal 1/6/2026 collegamento alla descrizione della causale contabile"
        )
        columns, unknown, _ = resolve_export_columns(prompt, "primanota", for_export=True)
        self.assertEqual(unknown, [])
        self.assertIn("Causale", columns)
        self.assertIn("DescrizioneCausaleContabile", columns)
        self.assertLess(
            columns.index("Causale"),
            columns.index("DescrizioneCausaleContabile"),
        )
        self.assertIn("DescrizioneDare", columns)
        self.assertIn("DescrizioneAvere", columns)
        result = ai_assistant._try_fast_path_table_export_sql(prompt)
        self.assertIsNotNone(result)
        sql, _ = result
        self.assertIn("LEFT JOIN causali_contabili cc", sql)
        self.assertIn('AS "DescrizioneCausaleContabile"', sql)
        self.assertIn("LEFT JOIN fornitori f_dare", sql)
        self.assertIn('p."Tipo" = 1', sql)

    def test_fast_path_primanota_export_with_conto_links(self):
        from apps.core.ai_export import (
            _extract_export_columns,
            enrich_primanota_export_columns,
        )

        prompt = (
            "export in xlsx tabella Primanota tipo generico "
            "numero registrazione, data registrazione, descrizione, "
            "contodare, dare, contoavere, avere, descrizione_avere "
            "dal 1/6/2026 a questa richiesta vorrei aggiungere un collegamento "
            "del contodare e contoavere con la seguente regola se il conto "
            "inizia per F deve cercare nei fornitori se inizia con C nei "
            "clienti negli altri casi nel PDC"
        )
        columns, unknown, _ = _extract_export_columns(prompt, "primanota")
        self.assertEqual(
            columns,
            [
                "NumeroReg",
                "DataReg",
                "Descrizione",
                "ContoDare",
                "Dare",
                "ContoAvere",
                "Avere",
                "DescrizioneAvere",
            ],
        )
        self.assertEqual(unknown, [])
        enriched = enrich_primanota_export_columns(prompt, columns)
        self.assertIn("DescrizioneDare", enriched)
        self.assertLess(
            enriched.index("ContoDare"),
            enriched.index("DescrizioneDare"),
        )
        result = ai_assistant._try_fast_path_table_export_sql(prompt)
        self.assertIsNotNone(result)
        sql, _ = result
        self.assertIn("LEFT JOIN fornitori f_dare", sql)
        self.assertIn("LEFT JOIN clienti c_dare", sql)
        self.assertIn("LEFT JOIN pdc pdc_dare", sql)
        self.assertIn("LEFT JOIN fornitori f_avere", sql)
        self.assertIn("LEFT JOIN clienti c_avere", sql)
        self.assertIn("LEFT JOIN pdc pdc_avere", sql)
        self.assertIn('AS "DescrizioneDare"', sql)
        self.assertIn('AS "DescrizioneAvere"', sql)
        self.assertIn("WHEN UPPER(LEFT(TRIM(BOTH FROM COALESCE(pd.\"ContoDare\"", sql)

    def test_fast_path_primanota_export_generico_dal_data(self):
        from apps.core.ai_export import (
            _extract_export_columns,
            detect_export_table_from_prompt,
            extract_primanota_export_filters,
        )

        prompt = (
            "export in xlsx tabella Primanota tipo generico "
            "contodare, dare, contoavere, avere dal 1/6/2026"
        )
        self.assertEqual(detect_export_table_from_prompt(prompt), "primanota")
        columns, unknown, _ = _extract_export_columns(prompt, "primanota")
        self.assertEqual(columns, ["ContoDare", "Dare", "ContoAvere", "Avere"])
        self.assertEqual(unknown, [])
        self.assertEqual(
            extract_primanota_export_filters(prompt),
            [
                'p."Tipo" = 1',
                "p.\"DataReg\" >= TIMESTAMP '2026-06-01'",
            ],
        )
        result = ai_assistant._try_fast_path_table_export_sql(prompt)
        self.assertIsNotNone(result)
        sql, spiegazione = result
        self.assertIn('pd."ContoDare"', sql)
        self.assertIn('pd."Dare"', sql)
        self.assertIn('pd."ContoAvere"', sql)
        self.assertIn('pd."Avere_Imponibile" AS "Avere"', sql)
        self.assertIn('p."Tipo" = 1', sql)
        self.assertIn("TIMESTAMP '2026-06-01'", sql)
        self.assertIn("JOIN primanota_dettaglio", sql)
        self.assertIn("Export tabella primanota", spiegazione)

    def test_fast_path_pdc_export_sql_exports_whole_table(self):
        prompt = (
            "crea un file dalla tabella PDC in formato xlsx con codice, descrizione"
        )
        # "descrizione" come colonna non deve diventare ricerca testuale
        self.assertIsNone(ai_assistant._extract_pdc_search_term(prompt))
        result = ai_assistant._try_fast_path_table_export_sql(prompt)
        self.assertIsNotNone(result)
        sql, spiegazione = result
        self.assertEqual(sql, 'SELECT "Codice", "Descrizione" FROM pdc')
        self.assertNotIn("ILIKE", sql)
        self.assertIn("Export tabella pdc", spiegazione)

    def test_fast_path_pdc_export_maps_tipo_labels(self):
        prompt = (
            "crea file xlsx della lista PDC con codice, descrizione, tipo "
            "sostituisci (tipo=2 come 'mastro', tipo=0 come 'conto', "
            "tipo=1 come 'sottoconto')"
        )
        from apps.core.ai_export import _extract_export_columns

        columns, unknown, _ = _extract_export_columns(prompt, "pdc")
        self.assertEqual(columns, ["Codice", "Descrizione", "Tipo"])
        self.assertEqual(unknown, [])
        result = ai_assistant._try_fast_path_table_export_sql(prompt)
        self.assertIsNotNone(result)
        sql, _ = result
        self.assertIn("CASE", sql)
        self.assertIn("WHEN 2 THEN 'mastro'", sql)
        self.assertIn("WHEN 0 THEN 'conto'", sql)
        self.assertIn("WHEN 1 THEN 'sottoconto'", sql)
        self.assertIn('END AS "Tipo"', sql)
        self.assertIn('"Codice"', sql)
        self.assertIn('"Descrizione"', sql)

    def test_fast_path_pdc_export_with_tipo_filter(self):
        from apps.core.ai_export import (
            _extract_export_columns,
            extract_export_where_clauses,
        )

        prompt = (
            "crea un file dalla tabella PDC in formato xlsx con codice, descrizione "
            "dove tipoconto = 1"
        )
        columns, unknown, _ = _extract_export_columns(prompt, "pdc")
        self.assertEqual(columns, ["Codice", "Descrizione"])
        self.assertEqual(unknown, [])
        self.assertEqual(
            extract_export_where_clauses(prompt, "pdc"),
            ['"Tipo" = 1'],
        )
        result = ai_assistant._try_fast_path_table_export_sql(prompt)
        self.assertIsNotNone(result)
        sql, spiegazione = result
        self.assertEqual(
            sql,
            'SELECT "Codice", "Descrizione" FROM pdc WHERE "Tipo" = 1',
        )
        self.assertIn('"Tipo" = 1', spiegazione)

    def test_fast_path_clienti_export_with_provincia_filter(self):
        from apps.core.ai_export import (
            _extract_export_columns,
            detect_export_table_from_prompt,
            extract_export_where_clauses,
        )

        prompt = (
            "crea file xlsx della lista Clienti con codice, ragione sociale "
            "dove provincia = LU"
        )
        self.assertEqual(detect_export_table_from_prompt(prompt), "clienti")
        columns, unknown, _ = _extract_export_columns(prompt, "clienti")
        self.assertEqual(columns, ["Codice", "RagioneSociale"])
        self.assertEqual(unknown, [])
        self.assertEqual(
            extract_export_where_clauses(prompt, "clienti"),
            ['UPPER("Provincia") = UPPER(\'LU\')'],
        )
        result = ai_assistant._try_fast_path_table_export_sql(prompt)
        self.assertIsNotNone(result)
        sql, _ = result
        self.assertIn("FROM clienti", sql)
        self.assertIn('AS "RagioneSociale"', sql)
        self.assertIn('UPPER("Provincia") = UPPER(\'LU\')', sql)
        self.assertNotIn("articoli", sql)

    def test_fast_path_clienti_export_joins_condizioni_for_descrizione_pagamento(self):
        from apps.core.ai_export import _extract_export_columns

        prompt = (
            "crea file xlsx della lista Clienti con codice, ragione sociale, "
            "provincia, cond. pagamento, descrizione_pagamento "
            "dove provincia = LU"
        )
        columns, unknown, _ = _extract_export_columns(prompt, "clienti")
        self.assertEqual(
            columns,
            [
                "Codice",
                "RagioneSociale",
                "Provincia",
                "CondPaga",
                "DescrizionePagamento",
            ],
        )
        self.assertEqual(unknown, [])
        result = ai_assistant._try_fast_path_table_export_sql(prompt)
        self.assertIsNotNone(result)
        sql, _ = result
        self.assertIn("LEFT JOIN condizioni", sql)
        self.assertIn(
            'UPPER(condizioni."Codice") = UPPER(clienti."CondPaga")',
            sql,
        )
        self.assertIn('condizioni."Descrizione" AS "DescrizionePagamento"', sql)
        self.assertIn('clienti."CondPaga"', sql)
        self.assertIn('UPPER(clienti."Provincia") = UPPER(\'LU\')', sql)

    @patch("apps.core.ai_assistant._get_client")
    @patch("apps.core.ai_assistant._call_llm")
    def test_ask_ai_clienti_export_uses_fast_path(
        self,
        call_llm_mock,
        get_client_mock,
    ):
        import tempfile
        from io import BytesIO
        from pathlib import Path

        from openpyxl import load_workbook

        prompt = (
            "crea file xlsx della lista Clienti con codice, ragione sociale "
            "dove provincia = LU"
        )
        rows = [
            {"Codice": "C1", "RagioneSociale": "Cliente Lucca"},
            {"Codice": "C2", "RagioneSociale": "Altro Lucca"},
        ]
        with tempfile.TemporaryDirectory() as tmp:
            with override_settings(AI_EXPORT_DIR=tmp):
                with patch(
                    "apps.core.ai_assistant._execute_query",
                    return_value=(rows, False),
                ):
                    result = ai_assistant.ask_ai(prompt, limit=10)

            self.assertTrue(result["ok"])
            self.assertTrue(result["fast_path"])
            self.assertTrue(result["export_requested"])
            self.assertEqual(result["export_headers"], ["Codice", "Ragione sociale"])
            call_llm_mock.assert_not_called()
            wb = load_workbook(
                BytesIO((Path(tmp) / f"{result['export_token']}.xlsx").read_bytes())
            )
            self.assertEqual(wb.active.max_row, 3)

    @patch("apps.core.ai_assistant._get_client")
    @patch("apps.core.ai_assistant._call_llm")
    def test_ask_ai_pdc_export_uses_fast_path_all_rows(
        self,
        call_llm_mock,
        get_client_mock,
    ):
        import tempfile
        from io import BytesIO
        from pathlib import Path

        from openpyxl import load_workbook

        prompt = (
            "crea un file dalla tabella PDC in formato xlsx con codice, descrizione"
        )
        rows = [
            {"Codice": "01", "Descrizione": "Attivo"},
            {"Codice": "01.01", "Descrizione": "Immobilizzazioni"},
            {"Codice": "02", "Descrizione": "Passivo"},
        ]
        with tempfile.TemporaryDirectory() as tmp:
            with override_settings(AI_EXPORT_DIR=tmp):
                with patch(
                    "apps.core.ai_assistant._execute_query",
                    return_value=(rows, False),
                ) as execute_mock:
                    result = ai_assistant.ask_ai(prompt, limit=10)

            self.assertTrue(result["ok"])
            self.assertTrue(result["fast_path"])
            self.assertTrue(result["export_requested"])
            self.assertEqual(result["conteggio"], 3)
            self.assertEqual(result["export_headers"], ["Codice", "Descrizione"])
            call_llm_mock.assert_not_called()
            sql = execute_mock.call_args.args[0]
            self.assertIn('FROM pdc', sql)
            self.assertNotIn("ILIKE", sql)

            token = result["export_token"]
            wb = load_workbook(BytesIO((Path(tmp) / f"{token}.xlsx").read_bytes()))
            self.assertEqual(wb.active.max_row, 4)

    def test_fast_path_sql_includes_categoria_join_for_descrizione_categoria(self):
        prompt = (
            "articoli con descrizione sinonimi di calzature crea un file xlsx con "
            "codice, categoria, descrizione_categoria"
        )
        result = ai_assistant._try_fast_path_articoli_sql(prompt)
        self.assertIsNotNone(result)
        sql, _ = result
        self.assertIn("LEFT JOIN categorie", sql)
        self.assertIn(
            'UPPER(categorie."Codice") = UPPER(articoli."CatOmogenea")',
            sql,
        )
        self.assertIn('categorie."Descrizione" AS "DescrizioneCategoria"', sql)
        self.assertIn('articoli."CatOmogenea"', sql)

    @patch("apps.core.ai_assistant._get_client")
    @patch("apps.core.ai_assistant._call_llm")
    def test_ask_ai_fornitore_export_uses_fast_path_with_valid_trim_sql(
        self,
        call_llm_mock,
        get_client_mock,
    ):
        rows = [
            {
                "Codice": "A1",
                "Descrizione": "Scarpe sportive",
                "CodFornitore": "F01",
                "RagioneSocialeFornitore": "Fornitore SpA",
            },
        ]
        with patch(
            "apps.core.ai_assistant._execute_query",
            return_value=(rows, False),
        ):
            result = ai_assistant.ask_ai(self.PROMPT_FORNITORE_EXPORT, limit=10)

        self.assertTrue(result["ok"])
        self.assertTrue(result["fast_path"])
        self.assertIsNotNone(result["sql"])
        self.assertNotIn('TRIM(BOTH,', result["sql"])
        self.assertIn('TRIM(BOTH FROM CONCAT(', result["sql"])
        self.assertIn('AS "RagioneSocialeFornitore"', result["sql"])
        call_llm_mock.assert_not_called()
        get_client_mock.assert_not_called()

    def test_detect_table_ignores_trim_from_concat_in_select(self):
        sql = (
            'SELECT articoli."Codice", TRIM(BOTH FROM CONCAT('
            'COALESCE(fornitori."RagioneSociale1", \'\'), \' \', '
            'COALESCE(fornitori."RagioneSociale2", \'\'))) AS "RagioneSocialeFornitore" '
            'FROM articoli LEFT JOIN fornitori ON UPPER(fornitori."Codice") = UPPER(articoli."CodFornitore")'
        )
        self.assertEqual(ai_assistant._detect_table(sql), "articoli")

    @patch("apps.core.ai_assistant._get_client")
    @patch("apps.core.ai_assistant._call_llm")
    def test_ask_ai_export_with_extra_field_in_xlsx(
        self,
        call_llm_mock,
        get_client_mock,
    ):
        import tempfile
        from io import BytesIO
        from pathlib import Path

        from openpyxl import load_workbook

        prompt = (
            "articoli con descrizione sinonimi di calzature "
            "genera xlsx con campi Codice, Descrizione, FlDisattivato"
        )
        rows = [
            {
                "Codice": "A1",
                "Descrizione": "Scarpe",
                "FlDisattivato": False,
                "Giacenza": True,
            },
        ]
        with tempfile.TemporaryDirectory() as tmp:
            with override_settings(AI_EXPORT_DIR=tmp):
                with patch(
                    "apps.core.ai_assistant._execute_query",
                    return_value=(rows, False),
                ):
                    result = ai_assistant.ask_ai(prompt, limit=10)

            self.assertTrue(result["ok"])
            self.assertTrue(result["fast_path"])
            self.assertTrue(result["export_requested"])
            self.assertIsNotNone(result["export_token"])
            self.assertEqual(len(result["export_headers"]), 3)
            self.assertTrue(
                any("disattiv" in str(h).lower() for h in result["export_headers"])
            )
            call_llm_mock.assert_not_called()

            token = result["export_token"]
            wb = load_workbook(BytesIO((Path(tmp) / f"{token}.xlsx").read_bytes()))
            ws = wb.active
            self.assertEqual(ws.max_column, 3)


class AiAssistantXlsxExportTests(SimpleTestCase):
    EXPORT_PROMPT = (
        "articoli con descrizione sinonimi di calzature genera un file xlsx"
    )

    def test_wants_xlsx_export_detects_keywords(self):
        self.assertTrue(ai_assistant._wants_xlsx_export(self.EXPORT_PROMPT))
        self.assertTrue(ai_assistant._wants_xlsx_export("esporta articoli in excel"))
        self.assertTrue(ai_assistant._wants_xlsx_export("scarica file excel articoli"))
        self.assertFalse(
            ai_assistant._wants_xlsx_export(
                "articoli con descrizione sinonimi di calzature"
            )
        )

    @patch("apps.core.ai_assistant._get_client")
    @patch("apps.core.ai_assistant._call_llm")
    def test_ask_ai_export_uses_fast_path_and_codice_descrizione(
        self,
        call_llm_mock,
        get_client_mock,
    ):
        import tempfile
        from io import BytesIO
        from pathlib import Path

        from openpyxl import load_workbook

        from apps.core import ai_export

        rows = [
            {"Codice": "A1", "Descrizione": "Scarpe running", "Giacenza": 3},
            {"Codice": "A2", "Descrizione": "Stivali"},
        ]
        with tempfile.TemporaryDirectory() as tmp:
            with override_settings(AI_EXPORT_DIR=tmp):
                with patch(
                    "apps.core.ai_assistant._execute_query",
                    return_value=(rows, False),
                ) as execute_mock:
                    result = ai_assistant.ask_ai(self.EXPORT_PROMPT, limit=10)

            self.assertTrue(result["ok"])
            self.assertTrue(result["fast_path"])
            self.assertTrue(result["export_requested"])
            self.assertIsNotNone(result["export_token"])
            self.assertEqual(result["export_headers"], ["Codice", "Descrizione"])
            self.assertTrue(str(result["download_filename"]).endswith(".xlsx"))
            call_llm_mock.assert_not_called()
            get_client_mock.assert_not_called()
            self.assertEqual(execute_mock.call_args.args[1], ai_assistant.EXPORT_MAX_ROWS)

            token = result["export_token"]
            xlsx_path = Path(tmp) / f"{token}.xlsx"
            self.assertTrue(xlsx_path.is_file())
            wb = load_workbook(BytesIO(xlsx_path.read_bytes()))
            ws = wb.active
            self.assertEqual([cell.value for cell in ws[1]], ["Codice", "Descrizione"])
            self.assertEqual([cell.value for cell in ws[2]], ["A1", "Scarpe running"])
            self.assertEqual([cell.value for cell in ws[3]], ["A2", "Stivali"])
            self.assertEqual(ws.max_column, 2)

            self.assertEqual(
                ai_export.select_export_columns(rows, "articoli"),
                ["Codice", "Descrizione"],
            )

    def test_select_export_columns_keeps_sql_columns_for_other_tables(self):
        from apps.core.ai_export import select_export_columns

        rows = [{"Codice": "1", "Descrizione": "Cassa", "TipoConto": "P"}]
        self.assertEqual(
            select_export_columns(rows, "pdc"),
            ["Codice", "Descrizione", "TipoConto"],
        )

    def test_select_export_columns_respects_requested_columns(self):
        from apps.core.ai_export import select_export_columns

        rows = [
            {"Codice": "A1", "Descrizione": "X", "Listino1": 10.0, "Giacenza": True},
        ]
        self.assertEqual(
            select_export_columns(
                rows, "articoli", requested_columns=["Codice", "Listino1"]
            ),
            ["Codice", "Listino1"],
        )


class AiAssistantCsvExportTests(SimpleTestCase):
    EXPORT_PROMPT = (
        "articoli con descrizione sinonimi di calzature crea un file csv con "
        "codice, descrizione, codice fornitore, ragione sociale fornitore, "
        "unitamisura, peso netto, peso lordo alias 'peso_lordo', categoria, "
        "descrizione_categoria, codicealternativo1 as 'Cod.Forn.Art.'"
    )

    @patch("apps.core.ai_assistant._get_client")
    @patch("apps.core.ai_assistant._call_llm")
    def test_ask_ai_export_creates_csv_with_header_aliases(
        self,
        call_llm_mock,
        get_client_mock,
    ):
        import csv
        import tempfile
        from io import StringIO
        from pathlib import Path

        rows = [
            {
                "Codice": "A1",
                "Descrizione": "Scarpe running",
                "CodFornitore": "F01",
                "RagioneSocialeFornitore": "Fornitore SpA",
                "UnitaMisura": "PA",
                "PesoNetto": 0.5,
                "PesoLordo_Manodopera": 0.7,
                "CatOmogenea": "CALZ",
                "DescrizioneCategoria": "Calzature",
                "CodiceAlternativo1": "FORN-99",
            },
        ]
        with tempfile.TemporaryDirectory() as tmp:
            with override_settings(AI_EXPORT_DIR=tmp):
                with patch(
                    "apps.core.ai_assistant._execute_query",
                    return_value=(rows, False),
                ):
                    result = ai_assistant.ask_ai(self.EXPORT_PROMPT, limit=10)

            self.assertTrue(result["ok"])
            self.assertTrue(result["fast_path"])
            self.assertTrue(result["export_requested"])
            self.assertIsNotNone(result["export_token"])
            self.assertTrue(str(result["download_filename"]).endswith(".csv"))
            self.assertIn("peso_lordo", result["export_headers"])
            self.assertIn("Cod.Forn.Art.", result["export_headers"])
            call_llm_mock.assert_not_called()

            token = result["export_token"]
            csv_path = Path(tmp) / f"{token}.csv"
            self.assertTrue(csv_path.is_file())
            text = csv_path.read_text(encoding="utf-8-sig")
            reader = csv.reader(StringIO(text), delimiter="\t")
            header = next(reader)
            self.assertEqual(header[-1], "Cod.Forn.Art.")
            self.assertIn("peso_lordo", header)
            data = next(reader)
            self.assertEqual(data[0], "A1")
            self.assertEqual(data[-1], "FORN-99")


class AiAssistantPrimaryKeyFixTests(SimpleTestCase):
    def test_removes_wrong_id_when_codice_present(self):
        sql = 'SELECT "ID", "Codice", "Descrizione" FROM articoli WHERE "Descrizione" ILIKE \'%calzature%\''
        fixed = ai_assistant._fix_primary_key_columns(sql)
        self.assertNotIn('"ID"', fixed)
        self.assertIn('"Codice"', fixed)

    def test_replaces_id_with_codice_when_codice_missing(self):
        sql = 'SELECT "ID", "Descrizione" FROM articoli LIMIT 3'
        fixed = ai_assistant._fix_primary_key_columns(sql)
        self.assertNotIn('"ID"', fixed)
        self.assertIn('"Codice"', fixed)

    def test_keeps_id_for_primanota(self):
        sql = 'SELECT "ID", "DataReg" FROM primanota LIMIT 10'
        fixed = ai_assistant._fix_primary_key_columns(sql)
        self.assertIn('"ID"', fixed)

    @override_settings(AI_BACKEND="openai")
    @patch("apps.core.ai_assistant._execute_query", return_value=([{"Codice": "A1", "Descrizione": "Calzature"}], False))
    @patch("apps.core.ai_assistant._get_model", return_value="test-model")
    @patch("apps.core.ai_assistant._get_client")
    def test_ask_ai_fixes_wrong_id_in_generated_sql(
        self,
        get_client_mock,
        _get_model_mock,
        _execute_query_mock,
    ):
        fake_client = SimpleNamespace()
        fake_client.chat = SimpleNamespace()
        fake_client.chat.completions = SimpleNamespace()
        fake_client.chat.completions.create = lambda **kwargs: SimpleNamespace(
            choices=[
                SimpleNamespace(
                    message=SimpleNamespace(
                        content=(
                            '{"sql":"SELECT \\"ID\\", \\"Codice\\", \\"Descrizione\\" FROM articoli '
                            'WHERE \\"Descrizione\\" ILIKE \'%calzature%\' LIMIT 3","spiegazione":"ok"}'
                        )
                    )
                )
            ]
        )
        get_client_mock.return_value = fake_client

        result = ai_assistant.ask_ai("articoli con calzature nella descrizione", limit=3)

        self.assertTrue(result["ok"])
        self.assertNotIn('"ID"', result["sql"])
        self.assertIn('"Codice"', result["sql"])
        execute_sql = _execute_query_mock.call_args[0][0]
        self.assertNotIn('"ID"', execute_sql)


class AiAssistantPrimanotaJoinFixTests(SimpleTestCase):
    def test_fixes_fk_on_primanota_alias(self):
        sql = (
            'SELECT p."ID", pd."Avere_Imponibile" FROM primanota p '
            'JOIN primanota_dettaglio pd ON p."ID" = p.id_added_by_converter '
            'WHERE p."Tipo" IN (2, 4)'
        )
        fixed = ai_assistant._fix_primanota_dettaglio_joins(sql)
        self.assertIn('p."ID" = pd."id_added_by_converter"', fixed)
        self.assertNotIn("p.id_added_by_converter", fixed)

    def test_fixes_join_on_detail_row_pk(self):
        sql = (
            'SELECT p."ID" FROM primanota p '
            'JOIN primanota_dettaglio pd ON p."ID" = pd."ID"'
        )
        fixed = ai_assistant._fix_primanota_dettaglio_joins(sql)
        self.assertIn('p."ID" = pd."id_added_by_converter"', fixed)

    def test_fixes_join_with_full_table_names(self):
        sql = (
            'SELECT primanota."ID" FROM primanota '
            'JOIN primanota_dettaglio ON primanota."ID" = primanota."id_added_by_converter"'
        )
        fixed = ai_assistant._fix_primanota_dettaglio_joins(sql)
        self.assertIn(
            'primanota."ID" = primanota_dettaglio."id_added_by_converter"', fixed
        )

    def test_quotes_id_added_by_converter(self):
        sql = (
            'SELECT p."ID" FROM primanota p '
            'JOIN primanota_dettaglio pd ON p."ID" = pd.id_added_by_converter'
        )
        fixed = ai_assistant._fix_column_quoting(sql)
        self.assertIn('pd."id_added_by_converter"', fixed)


class AiAssistantPrimanotaListRouteTests(SimpleTestCase):
    def test_primanota_in_table_list_routes(self):
        self.assertIn("primanota", ai_assistant.TABLE_LIST_ROUTES)
        self.assertEqual(
            ai_assistant.TABLE_LIST_ROUTES["primanota"],
            "primanota:list",
        )

    def test_resolve_ai_list_table_maps_dettaglio_to_testa(self):
        self.assertEqual(
            ai_assistant.resolve_ai_list_table("primanota_dettaglio"),
            "primanota",
        )
        self.assertEqual(
            ai_assistant.resolve_ai_list_table("primanota"),
            "primanota",
        )

    @patch("apps.core.ai_assistant._get_client")
    @patch("apps.core.ai_assistant._call_llm")
    def test_ask_ai_fast_path_sets_table_for_list_redirect(
        self,
        call_llm_mock,
        get_client_mock,
    ):
        prompt = (
            "Cerca in Primanota IVA dove imponibile è compreso tra 1500 e 1750 "
            "euro nell'anno in corso"
        )
        with patch(
            "apps.core.ai_assistant._execute_query",
            return_value=([{"ID": 1, "Avere_Imponibile": 1600}], False),
        ):
            result = ai_assistant.ask_ai(prompt, limit=10)
        self.assertTrue(result["ok"])
        self.assertEqual(result["table"], "primanota")
        self.assertIsNotNone(result["link"])
        self.assertEqual(result["link"]["pk_column"], "ID")
        call_llm_mock.assert_not_called()
        get_client_mock.assert_not_called()


class AiAssistantPrimanotaFastPathTests(SimpleTestCase):
    def test_detects_primanota_iva_imponibile_request(self):
        self.assertTrue(
            ai_assistant._is_primanota_iva_imponibile_request(
                "Cerca in Primanota IVA dove imponibile è compreso tra 1500 e 1750 euro nell'anno in corso"
            )
        )

    def test_extracts_imponibile_range(self):
        amount_range = ai_assistant._extract_imponibile_range(
            "imponibile compreso tra 1500 e 1750 euro"
        )
        self.assertEqual(amount_range, (1500.0, 1750.0))

    def test_fast_path_generates_correct_join_sql(self):
        prompt = (
            "Cerca in Primanota IVA dove imponibile è compreso tra 1500 e 1750 "
            "euro nell'anno in corso"
        )
        result = ai_assistant._try_fast_path_primanota_iva_sql(prompt)
        self.assertIsNotNone(result)
        sql, spiegazione = result
        self.assertIn('p."ID" = pd."id_added_by_converter"', sql)
        self.assertIn('pd."Avere_Imponibile" BETWEEN 1500 AND 1750', sql)
        self.assertIn('p."Tipo" IN (2, 4)', sql)
        self.assertIn('EXTRACT(YEAR FROM p."DataReg")', sql)
        self.assertIn("anno in corso", spiegazione.lower())

    def test_build_user_prompt_adds_primanota_join_hint(self):
        prompt = (
            "mostrami primanota iva con imponibile superiore a 1000 nell'anno in corso"
        )
        guarded = ai_assistant._build_ai_user_prompt(prompt)
        self.assertIn('primanota."ID" = primanota_dettaglio."id_added_by_converter"', guarded)
        self.assertIn("id_added_by_converter", guarded)

    @patch("apps.core.ai_assistant._get_client")
    @patch("apps.core.ai_assistant._call_llm")
    def test_ask_ai_uses_primanota_fast_path_without_llm(
        self,
        call_llm_mock,
        get_client_mock,
    ):
        prompt = (
            "Cerca in Primanota IVA dove imponibile è compreso tra 1500 e 1750 "
            "euro nell'anno in corso"
        )
        with patch(
            "apps.core.ai_assistant._execute_query",
            return_value=([{"ID": 1, "Avere_Imponibile": 1600}], False),
        ):
            result = ai_assistant.ask_ai(prompt, limit=10)
        self.assertTrue(result["ok"])
        self.assertTrue(result["fast_path"])
        self.assertIn('pd."id_added_by_converter"', result["sql"])
        call_llm_mock.assert_not_called()
        get_client_mock.assert_not_called()

    @override_settings(AI_BACKEND="openai")
    @patch("apps.core.ai_assistant._try_fast_path_sql", return_value=None)
    @patch("apps.core.ai_assistant._execute_query", return_value=([], False))
    @patch("apps.core.ai_assistant._get_model", return_value="test-model")
    @patch("apps.core.ai_assistant._get_client")
    def test_ask_ai_fixes_wrong_primanota_join_from_llm(
        self,
        get_client_mock,
        _get_model_mock,
        _execute_query_mock,
        _fast_path_mock,
    ):
        fake_client = SimpleNamespace()
        fake_client.chat = SimpleNamespace()
        fake_client.chat.completions = SimpleNamespace()
        fake_client.chat.completions.create = lambda **kwargs: SimpleNamespace(
            choices=[
                SimpleNamespace(
                    message=SimpleNamespace(
                        content=(
                            '{"sql":"SELECT p.\\"ID\\", pd.\\"Avere_Imponibile\\" FROM primanota p '
                            'JOIN primanota_dettaglio pd ON p.\\"ID\\" = p.id_added_by_converter '
                            'WHERE p.\\"Tipo\\" IN (2, 4) LIMIT 10","spiegazione":"ok"}'
                        )
                    )
                )
            ]
        )
        get_client_mock.return_value = fake_client

        result = ai_assistant.ask_ai(
            "primanota iva imponibile tra 1500 e 1750", limit=10
        )

        self.assertTrue(result["ok"])
        self.assertIn('p."ID" = pd."id_added_by_converter"', result["sql"])
        execute_sql = _execute_query_mock.call_args[0][0]
        self.assertIn('pd."id_added_by_converter"', execute_sql)
        self.assertNotIn("p.id_added_by_converter", execute_sql)


class AiAssistantPrimanotaTotaleDocumentoFastPathTests(SimpleTestCase):
    PROMPT = (
        "Cerca in Primanota IVA dove totaledocumento è compreso tra 1500 e 1750 "
        "euro nell'anno in corso"
    )

    def test_detects_primanota_iva_totale_documento_request(self):
        self.assertTrue(
            ai_assistant._is_primanota_iva_totale_documento_request(self.PROMPT)
        )
        self.assertFalse(
            ai_assistant._is_primanota_iva_imponibile_request(self.PROMPT)
        )

    def test_fast_path_generates_grouped_having_sql(self):
        result = ai_assistant._try_fast_path_primanota_iva_totale_documento_sql(
            self.PROMPT
        )
        self.assertIsNotNone(result)
        sql, spiegazione = result
        self.assertIn('p."ID" = pd."id_added_by_converter"', sql)
        self.assertIn('p."Tipo" IN (2, 4)', sql)
        self.assertIn('pd."dummy" IS NOT TRUE', sql)
        self.assertIn('GROUP BY p."ID"', sql)
        self.assertIn(
            'HAVING SUM(COALESCE(pd."Avere_Imponibile", 0) + '
            'COALESCE(pd."ImportoIva", 0)) BETWEEN 1500 AND 1750',
            sql,
        )
        self.assertIn('EXTRACT(YEAR FROM p."DataReg")', sql)
        self.assertIn("totale documento", spiegazione.lower())
        self.assertIn("anno in corso", spiegazione.lower())

    def test_build_user_prompt_adds_totale_documento_hint(self):
        guarded = ai_assistant._build_ai_user_prompt(self.PROMPT)
        self.assertIn('GROUP BY primanota."ID"', guarded)
        self.assertIn("HAVING SUM", guarded)
        self.assertIn("dummy", guarded)

    @patch("apps.core.ai_assistant._get_client")
    @patch("apps.core.ai_assistant._call_llm")
    def test_ask_ai_uses_totale_documento_fast_path_without_llm(
        self,
        call_llm_mock,
        get_client_mock,
    ):
        with patch(
            "apps.core.ai_assistant._execute_query",
            return_value=([{"ID": 42, "DataReg": "01/01/2026"}], False),
        ):
            result = ai_assistant.ask_ai(self.PROMPT, limit=10)
        self.assertTrue(result["ok"])
        self.assertTrue(result["fast_path"])
        self.assertEqual(result["table"], "primanota")
        self.assertEqual(result["link"]["pk_column"], "ID")
        self.assertEqual(ai_assistant.resolve_ai_list_table(result["table"]), "primanota")
        self.assertIn("GROUP BY", result["sql"])
        self.assertIn("HAVING SUM", result["sql"])
        call_llm_mock.assert_not_called()
        get_client_mock.assert_not_called()

    def test_totale_documento_takes_precedence_over_imponibile(self):
        prompt = (
            "primanota iva con totale documento e imponibile compreso tra 100 e 200"
        )
        result = ai_assistant._try_fast_path_sql(prompt)
        self.assertIsNotNone(result)
        sql, _ = result
        self.assertIn("HAVING SUM", sql)
        self.assertNotIn('pd."Avere_Imponibile" BETWEEN', sql)


class AiAssistantRateLimitTests(SimpleTestCase):
    @override_settings(AI_BACKEND="groq")
    @patch("apps.core.ai_assistant._is_ollama_available", return_value=False)
    @patch("apps.core.ai_assistant._get_client")
    def test_ask_ai_returns_friendly_message_on_groq_429(
        self,
        get_client_mock,
        _ollama_mock,
    ):
        class RateLimitError(Exception):
            pass

        fake_client = SimpleNamespace()
        fake_client.chat = SimpleNamespace()
        fake_client.chat.completions = SimpleNamespace()

        def fake_create(**kwargs):
            raise RateLimitError(
                "Error code: 429 - Rate limit reached for model "
                "openai/gpt-oss-20b ... requests per minute (RPM): Limit 30"
            )

        fake_client.chat.completions.create = fake_create
        get_client_mock.return_value = fake_client

        with patch(
            "apps.core.ai_assistant._is_groq_rate_limit_error",
            side_effect=lambda exc: "429" in str(exc),
        ):
            result = ai_assistant.ask_ai("mostrami i clienti")

        self.assertFalse(result["ok"])
        self.assertIn("Limite Groq raggiunto", result["errore"])
        self.assertNotIn("Error code: 429", result["errore"])

    @override_settings(AI_BACKEND="groq")
    @patch("apps.core.ai_assistant._is_ollama_available", return_value=False)
    @patch("apps.core.ai_assistant._get_client")
    def test_ask_ai_returns_tpd_message_when_ollama_unavailable(
        self,
        get_client_mock,
        _ollama_mock,
    ):
        fake_client = SimpleNamespace()
        fake_client.chat = SimpleNamespace()
        fake_client.chat.completions = SimpleNamespace()
        fake_client.chat.completions.create = lambda **kwargs: (_ for _ in ()).throw(
            Exception("Error code: 429 - tokens per day (TPD): Limit 200000")
        )
        get_client_mock.return_value = fake_client

        with patch(
            "apps.core.ai_assistant._is_groq_rate_limit_error",
            side_effect=lambda exc: "429" in str(exc),
        ), patch(
            "apps.core.ai_assistant._is_groq_tpd_limit_error",
            side_effect=lambda exc: "tokens per day" in str(exc).lower(),
        ):
            result = ai_assistant.ask_ai("mostrami i clienti")

        self.assertFalse(result["ok"])
        self.assertIn("Limite giornaliero Groq esaurito", result["errore"])

    @override_settings(AI_BACKEND="groq", OLLAMA_MODEL="qwen2.5:3b")
    @patch("apps.core.ai_assistant._execute_query", return_value=([], False))
    @patch("apps.core.ai_assistant._is_ollama_available", return_value=True)
    @patch("apps.core.ai_assistant._get_ollama_client")
    @patch("apps.core.ai_assistant._get_client")
    def test_ask_ai_falls_back_to_ollama_on_groq_tpd_limit(
        self,
        get_client_mock,
        get_ollama_client_mock,
        _ollama_available_mock,
        _execute_query_mock,
    ):
        groq_client = SimpleNamespace()
        groq_client.chat = SimpleNamespace()
        groq_client.chat.completions = SimpleNamespace()
        groq_client.chat.completions.create = lambda **kwargs: (_ for _ in ()).throw(
            Exception("Error code: 429 - tokens per day (TPD): Limit 200000")
        )
        get_client_mock.return_value = groq_client

        ollama_client = SimpleNamespace()
        ollama_client.chat = SimpleNamespace()
        ollama_client.chat.completions = SimpleNamespace()
        ollama_client.chat.completions.create = lambda **kwargs: SimpleNamespace(
            choices=[
                SimpleNamespace(
                    message=SimpleNamespace(
                        content='{"sql":"SELECT \\"Codice\\" FROM clienti LIMIT 1","spiegazione":"ok"}'
                    )
                )
            ]
        )
        get_ollama_client_mock.return_value = ollama_client

        with patch(
            "apps.core.ai_assistant._is_groq_rate_limit_error",
            side_effect=lambda exc: "429" in str(exc),
        ), patch(
            "apps.core.ai_assistant._is_groq_tpd_limit_error",
            side_effect=lambda exc: "tokens per day" in str(exc).lower(),
        ):
            result = ai_assistant.ask_ai("mostrami i clienti")

        self.assertTrue(result["ok"])
        get_ollama_client_mock.assert_called_once()

    @override_settings(AI_BACKEND="groq", OLLAMA_MODEL="llama3.1")
    @patch("apps.core.ai_assistant._execute_query", return_value=([], False))
    @patch("apps.core.ai_assistant._is_ollama_available", return_value=True)
    @patch("apps.core.ai_assistant._get_ollama_client")
    @patch("apps.core.ai_assistant._get_client")
    def test_ask_ai_falls_back_to_ollama_on_groq_429(
        self,
        get_client_mock,
        get_ollama_client_mock,
        _ollama_available_mock,
        _execute_query_mock,
    ):
        groq_client = SimpleNamespace()
        groq_client.chat = SimpleNamespace()
        groq_client.chat.completions = SimpleNamespace()
        groq_client.chat.completions.create = lambda **kwargs: (_ for _ in ()).throw(
            Exception("Error code: 429 - Rate limit reached")
        )
        get_client_mock.return_value = groq_client

        ollama_client = SimpleNamespace()
        ollama_client.chat = SimpleNamespace()
        ollama_client.chat.completions = SimpleNamespace()
        ollama_client.chat.completions.create = lambda **kwargs: SimpleNamespace(
            choices=[
                SimpleNamespace(
                    message=SimpleNamespace(
                        content='{"sql":"SELECT \\"Codice\\" FROM clienti LIMIT 1","spiegazione":"ok"}'
                    )
                )
            ]
        )
        get_ollama_client_mock.return_value = ollama_client

        result = ai_assistant.ask_ai("mostrami i clienti")

        self.assertTrue(result["ok"])
        get_ollama_client_mock.assert_called_once()


class AiAssistantPdcFastPathTests(SimpleTestCase):
    PROMPT = "cerca nel Piano dei conti dove descrizione è cassa"

    def test_detects_pdc_description_search(self):
        self.assertTrue(ai_assistant._is_explicit_pdc_text_search(self.PROMPT))

    def test_extracts_pdc_search_term(self):
        term = ai_assistant._extract_pdc_search_term(self.PROMPT)
        self.assertEqual(term, "cassa")

    def test_fast_path_generates_pdc_ilike_sql(self):
        result = ai_assistant._try_fast_path_pdc_sql(self.PROMPT)
        self.assertIsNotNone(result)
        sql, spiegazione = result
        self.assertIn("FROM pdc", sql)
        self.assertIn('"Descrizione" ILIKE \'%cassa%\'', sql)
        self.assertIn("Codice", sql)
        self.assertIn("cassa", spiegazione.lower())

    def test_pdc_in_table_list_routes(self):
        self.assertIn("pdc", ai_assistant.TABLE_LIST_ROUTES)
        self.assertEqual(ai_assistant.TABLE_LIST_ROUTES["pdc"], "pdc:list")

    def test_build_user_prompt_adds_pdc_description_hint(self):
        guarded = ai_assistant._build_ai_user_prompt(self.PROMPT)
        self.assertIn('pdc."Descrizione"', guarded)
        self.assertIn("ILIKE", guarded)

    @patch("apps.core.ai_assistant._get_client")
    @patch("apps.core.ai_assistant._call_llm")
    def test_ask_ai_uses_pdc_fast_path_without_llm(
        self,
        call_llm_mock,
        get_client_mock,
    ):
        with patch(
            "apps.core.ai_assistant._execute_query",
            return_value=([{"Codice": "4.01", "Descrizione": "Cassa"}], False),
        ):
            result = ai_assistant.ask_ai(self.PROMPT, limit=10)
        self.assertTrue(result["ok"])
        self.assertTrue(result["fast_path"])
        self.assertEqual(result["table"], "pdc")
        self.assertEqual(result["link"]["pk_column"], "Codice")
        self.assertEqual(ai_assistant.resolve_ai_list_table(result["table"]), "pdc")
        self.assertIn('"Descrizione" ILIKE \'%cassa%\'', result["sql"])
        call_llm_mock.assert_not_called()
        get_client_mock.assert_not_called()


class AiAssistantChartTests(SimpleTestCase):
    CHART_PROMPT = (
        "Cerca in Primanota IVA dove imponibile è compreso tra 15000 e 17500 euro "
        "nell'anno in corso e crea un grafico per mese di registrazione"
    )
    PIE_CHART_PROMPT = (
        "Cerca in Primanota IVA dove imponibile è compreso tra 15000 e 17500 euro "
        "nell'anno in corso e crea un grafico a torta per mese di registrazione"
    )
    PIE_CHART_WITH_PERCENTAGES_PROMPT = (
        "Cerca in Primanota IVA dove imponibile è compreso tra 15000 e 17500 euro "
        "nell'anno in corso e crea un grafico a torta per mese di registrazione "
        "aggiungi le percentuali"
    )
    LINE_CHART_PROMPT = (
        "Cerca in Primanota IVA dove imponibile è compreso tra 15000 e 17500 euro "
        "nell'anno in corso e crea un grafico a linee per mese di registrazione"
    )
    AREA_CHART_PROMPT = (
        "Cerca in Primanota IVA dove imponibile è compreso tra 15000 e 17500 euro "
        "nell'anno in corso e crea un grafico ad area per mese di registrazione"
    )
    HORIZONTAL_BAR_CHART_PROMPT = (
        "Cerca in Primanota IVA dove imponibile è compreso tra 15000 e 17500 euro "
        "nell'anno in corso e crea un grafico a barre orizzontali per mese di registrazione"
    )
    RADAR_CHART_PROMPT = (
        "Cerca in Primanota IVA dove imponibile è compreso tra 15000 e 17500 euro "
        "nell'anno in corso e crea un grafico radar per mese di registrazione"
    )

    def test_wants_percentages_detects_percentuali_keyword(self):
        self.assertTrue(
            ai_assistant._wants_percentages(
                "grafico a torta con percentuali per mese"
            )
        )

    def test_wants_percentages_detects_percent_sign(self):
        self.assertTrue(ai_assistant._wants_percentages("mostra distribuzione 50%"))

    def test_wants_percentages_false_without_keywords(self):
        self.assertFalse(ai_assistant._wants_percentages(self.CHART_PROMPT))

    def test_monthly_chart_percentages_rounds_values(self):
        self.assertEqual(
            ai_assistant._monthly_chart_percentages([3, 1]),
            [75, 25],
        )
        self.assertEqual(ai_assistant._monthly_chart_percentages([]), [])
        self.assertEqual(ai_assistant._monthly_chart_percentages([0, 0]), [0, 0])

    def test_wants_chart_detects_grafico_keyword(self):
        self.assertTrue(ai_assistant._wants_chart(self.CHART_PROMPT))

    def test_wants_chart_detects_per_mese_without_grafico(self):
        self.assertTrue(
            ai_assistant._wants_chart(
                "primanota iva imponibile tra 1500 e 1750 per mese"
            )
        )

    def test_wants_pie_chart_detects_torta_keyword(self):
        self.assertTrue(ai_assistant._wants_pie_chart(self.PIE_CHART_PROMPT))

    def test_wants_pie_chart_detects_pie_keyword(self):
        self.assertTrue(
            ai_assistant._wants_pie_chart(
                "primanota iva imponibile tra 1500 e 1750 pie per mese"
            )
        )

    def test_resolve_chart_type_defaults_to_bar(self):
        self.assertEqual(ai_assistant._resolve_chart_type(self.CHART_PROMPT), "bar")

    def test_resolve_chart_type_returns_pie_for_torta(self):
        self.assertEqual(ai_assistant._resolve_chart_type(self.PIE_CHART_PROMPT), "pie")

    def test_wants_line_chart_detects_andamento_keyword(self):
        self.assertTrue(ai_assistant._wants_line_chart(
            "primanota iva imponibile tra 1500 e 1750 andamento per mese"
        ))

    def test_wants_line_chart_detects_trend_keyword(self):
        self.assertTrue(ai_assistant._wants_line_chart(
            "mostra trend mensile primanota iva imponibile tra 1500 e 1750"
        ))

    def test_wants_area_chart_detects_grafico_ad_area(self):
        self.assertTrue(ai_assistant._wants_area_chart(self.AREA_CHART_PROMPT))

    def test_wants_area_chart_detects_area_with_grafico_context(self):
        self.assertTrue(ai_assistant._wants_area_chart(
            "grafico area primanota iva imponibile tra 1500 e 1750 per mese"
        ))

    def test_wants_horizontal_bar_chart_detects_barre_orizzontali(self):
        self.assertTrue(
            ai_assistant._wants_horizontal_bar_chart(self.HORIZONTAL_BAR_CHART_PROMPT)
        )

    def test_wants_radar_chart_detects_ragnatela(self):
        self.assertTrue(ai_assistant._wants_radar_chart(
            "primanota iva imponibile tra 1500 e 1750 ragnatela per mese"
        ))

    def test_resolve_chart_type_returns_line(self):
        self.assertEqual(
            ai_assistant._resolve_chart_type(self.LINE_CHART_PROMPT), "line"
        )

    def test_resolve_chart_type_returns_area(self):
        self.assertEqual(
            ai_assistant._resolve_chart_type(self.AREA_CHART_PROMPT), "area"
        )

    def test_resolve_chart_type_returns_horizontal_bar(self):
        self.assertEqual(
            ai_assistant._resolve_chart_type(self.HORIZONTAL_BAR_CHART_PROMPT),
            "horizontalBar",
        )

    def test_resolve_chart_type_returns_radar(self):
        self.assertEqual(
            ai_assistant._resolve_chart_type(self.RADAR_CHART_PROMPT), "radar"
        )

    def test_resolve_chart_type_from_prompt_alias(self):
        self.assertEqual(
            ai_assistant._resolve_chart_type_from_prompt(self.LINE_CHART_PROMPT),
            "line",
        )

    def test_monthly_chart_sql_aggregates_by_datareg_month(self):
        sql = ai_assistant._build_primanota_iva_imponibile_monthly_chart_sql(
            15000, 17500, True
        )
        self.assertIn('EXTRACT(MONTH FROM p."DataReg")', sql)
        self.assertIn('pd."Avere_Imponibile" BETWEEN 15000 AND 17500', sql)
        self.assertIn("GROUP BY", sql)
        self.assertIn('p."Tipo" IN (2, 4)', sql)

    def test_execute_monthly_bar_chart_fills_all_months(self):
        mock_cursor = MagicMock()
        mock_cursor.fetchall.return_value = [(3, 5), (8, 2)]
        mock_connection = MagicMock()
        mock_connection.cursor.return_value.__enter__.return_value = mock_cursor
        with patch("apps.core.ai_assistant.connection", mock_connection):
            chart = ai_assistant._execute_monthly_bar_chart(
                "SELECT 1",
                title="Test",
                dataset_label="Registrazioni",
            )
        self.assertEqual(chart["type"], "bar")
        self.assertEqual(len(chart["labels"]), 12)
        self.assertEqual(chart["datasets"][0]["data"][2], 5)
        self.assertEqual(chart["datasets"][0]["data"][7], 2)
        self.assertEqual(chart["datasets"][0]["data"][0], 0)
        self.assertNotIn("showPercentages", chart)

    def test_execute_monthly_pie_chart_filters_zero_months(self):
        mock_cursor = MagicMock()
        mock_cursor.fetchall.return_value = [(3, 5), (8, 2)]
        mock_connection = MagicMock()
        mock_connection.cursor.return_value.__enter__.return_value = mock_cursor
        with patch("apps.core.ai_assistant.connection", mock_connection):
            chart = ai_assistant._execute_monthly_chart(
                "SELECT 1",
                title="Test",
                dataset_label="Registrazioni",
                chart_type="pie",
            )
        self.assertEqual(chart["type"], "pie")
        self.assertEqual(chart["labels"], ["Mar", "Ago"])
        self.assertEqual(chart["datasets"][0]["data"], [5, 2])
        self.assertTrue(chart["showPercentages"])
        self.assertEqual(chart["datasets"][0]["percentages"], [71, 29])

    def test_execute_monthly_line_chart_uses_all_months(self):
        mock_cursor = MagicMock()
        mock_cursor.fetchall.return_value = [(3, 5), (8, 2)]
        mock_connection = MagicMock()
        mock_connection.cursor.return_value.__enter__.return_value = mock_cursor
        with patch("apps.core.ai_assistant.connection", mock_connection):
            chart = ai_assistant._execute_monthly_chart(
                "SELECT 1",
                title="Test",
                dataset_label="Registrazioni",
                chart_type="line",
            )
        self.assertEqual(chart["type"], "line")
        self.assertEqual(len(chart["labels"]), 12)
        self.assertEqual(chart["datasets"][0]["data"][2], 5)

    def test_execute_monthly_area_chart_uses_all_months(self):
        mock_cursor = MagicMock()
        mock_cursor.fetchall.return_value = [(1, 4)]
        mock_connection = MagicMock()
        mock_connection.cursor.return_value.__enter__.return_value = mock_cursor
        with patch("apps.core.ai_assistant.connection", mock_connection):
            chart = ai_assistant._execute_monthly_chart(
                "SELECT 1",
                title="Test",
                dataset_label="Registrazioni",
                chart_type="area",
            )
        self.assertEqual(chart["type"], "area")
        self.assertEqual(chart["datasets"][0]["data"][0], 4)

    def test_execute_monthly_horizontal_bar_chart(self):
        mock_cursor = MagicMock()
        mock_cursor.fetchall.return_value = [(6, 7)]
        mock_connection = MagicMock()
        mock_connection.cursor.return_value.__enter__.return_value = mock_cursor
        with patch("apps.core.ai_assistant.connection", mock_connection):
            chart = ai_assistant._execute_monthly_chart(
                "SELECT 1",
                title="Test",
                dataset_label="Registrazioni",
                chart_type="horizontalBar",
            )
        self.assertEqual(chart["type"], "horizontalBar")
        self.assertEqual(chart["datasets"][0]["data"][5], 7)

    def test_execute_monthly_radar_chart(self):
        mock_cursor = MagicMock()
        mock_cursor.fetchall.return_value = [(12, 3)]
        mock_connection = MagicMock()
        mock_connection.cursor.return_value.__enter__.return_value = mock_cursor
        with patch("apps.core.ai_assistant.connection", mock_connection):
            chart = ai_assistant._execute_monthly_chart(
                "SELECT 1",
                title="Test",
                dataset_label="Registrazioni",
                chart_type="radar",
            )
        self.assertEqual(chart["type"], "radar")
        self.assertEqual(chart["datasets"][0]["data"][11], 3)

    def test_try_build_primanota_iva_imponibile_line_chart(self):
        mock_cursor = MagicMock()
        mock_cursor.fetchall.return_value = [(1, 3), (2, 1)]
        mock_connection = MagicMock()
        mock_connection.cursor.return_value.__enter__.return_value = mock_cursor
        with patch("apps.core.ai_assistant.connection", mock_connection):
            chart = ai_assistant._try_build_primanota_iva_imponibile_chart(
                self.LINE_CHART_PROMPT
            )
        self.assertIsNotNone(chart)
        self.assertEqual(chart["type"], "line")
        self.assertEqual(len(chart["labels"]), 12)

    def test_try_build_primanota_iva_imponibile_radar_chart(self):
        mock_cursor = MagicMock()
        mock_cursor.fetchall.return_value = [(1, 2), (4, 1)]
        mock_connection = MagicMock()
        mock_connection.cursor.return_value.__enter__.return_value = mock_cursor
        with patch("apps.core.ai_assistant.connection", mock_connection):
            chart = ai_assistant._try_build_primanota_iva_imponibile_chart(
                self.RADAR_CHART_PROMPT
            )
        self.assertIsNotNone(chart)
        self.assertEqual(chart["type"], "radar")

    def test_try_build_chart_requires_monthly_intent(self):
        prompt = (
            "Cerca in Primanota IVA dove imponibile è compreso tra 15000 e 17500 "
            "euro nell'anno in corso e crea un grafico"
        )
        self.assertIsNone(ai_assistant._try_build_primanota_iva_imponibile_chart(prompt))

    def test_try_build_primanota_iva_imponibile_chart(self):
        mock_cursor = MagicMock()
        mock_cursor.fetchall.return_value = [(1, 3), (2, 1)]
        mock_connection = MagicMock()
        mock_connection.cursor.return_value.__enter__.return_value = mock_cursor
        with patch("apps.core.ai_assistant.connection", mock_connection):
            chart = ai_assistant._try_build_primanota_iva_imponibile_chart(self.CHART_PROMPT)
        self.assertIsNotNone(chart)
        self.assertEqual(chart["datasets"][0]["data"][0], 3)
        self.assertEqual(chart["datasets"][0]["data"][1], 1)
        self.assertIn("15000", chart["title"])

    def test_try_build_primanota_iva_imponibile_pie_chart(self):
        mock_cursor = MagicMock()
        mock_cursor.fetchall.return_value = [(1, 3), (2, 1), (5, 0)]
        mock_connection = MagicMock()
        mock_connection.cursor.return_value.__enter__.return_value = mock_cursor
        with patch("apps.core.ai_assistant.connection", mock_connection):
            chart = ai_assistant._try_build_primanota_iva_imponibile_chart(
                self.PIE_CHART_PROMPT
            )
        self.assertIsNotNone(chart)
        self.assertEqual(chart["type"], "pie")
        self.assertEqual(chart["labels"], ["Gen", "Feb"])
        self.assertEqual(chart["datasets"][0]["data"], [3, 1])
        self.assertTrue(chart["showPercentages"])
        self.assertEqual(chart["datasets"][0]["percentages"], [75, 25])

    def test_try_build_primanota_iva_imponibile_pie_chart_with_percentages_prompt(self):
        mock_cursor = MagicMock()
        mock_cursor.fetchall.return_value = [(1, 3), (2, 1)]
        mock_connection = MagicMock()
        mock_connection.cursor.return_value.__enter__.return_value = mock_cursor
        with patch("apps.core.ai_assistant.connection", mock_connection):
            chart = ai_assistant._try_build_primanota_iva_imponibile_chart(
                self.PIE_CHART_WITH_PERCENTAGES_PROMPT
            )
        self.assertIsNotNone(chart)
        self.assertEqual(chart["type"], "pie")
        self.assertTrue(chart["showPercentages"])
        self.assertEqual(chart["datasets"][0]["percentages"], [75, 25])

    def test_ask_ai_returns_chart_for_primanota_imponibile_prompt(self):
        mock_cursor = MagicMock()
        mock_cursor.description = [
            SimpleNamespace(name="ID"),
            SimpleNamespace(name="DataReg"),
            SimpleNamespace(name="Avere_Imponibile"),
        ]
        mock_cursor.fetchmany.return_value = [(1, "2026-01-15", 16000)]
        mock_cursor.fetchall.side_effect = [
            [(1, 2), (3, 1)],
            [(1, 2), (3, 1)],
        ]
        mock_connection = MagicMock()
        mock_connection.cursor.return_value.__enter__.return_value = mock_cursor
        with (
            patch("apps.core.ai_assistant._get_client") as get_client_mock,
            patch("apps.core.ai_assistant._call_llm") as call_llm_mock,
            patch("apps.core.ai_assistant.connection", mock_connection),
        ):
            result = ai_assistant.ask_ai(self.CHART_PROMPT, limit=10)

        self.assertTrue(result["ok"])
        self.assertTrue(result["chart_requested"])
        self.assertIsNotNone(result["chart"])
        self.assertEqual(result["chart"]["type"], "bar")
        self.assertEqual(len(result["chart"]["labels"]), 12)
        call_llm_mock.assert_not_called()
        get_client_mock.assert_not_called()

    def test_ask_ai_returns_pie_chart_for_torta_prompt(self):
        mock_cursor = MagicMock()
        mock_cursor.description = [
            SimpleNamespace(name="ID"),
            SimpleNamespace(name="DataReg"),
            SimpleNamespace(name="Avere_Imponibile"),
        ]
        mock_cursor.fetchmany.return_value = [(1, "2026-01-15", 16000)]
        mock_cursor.fetchall.side_effect = [
            [(1, 2), (3, 1)],
            [(1, 2), (3, 1)],
        ]
        mock_connection = MagicMock()
        mock_connection.cursor.return_value.__enter__.return_value = mock_cursor
        with (
            patch("apps.core.ai_assistant._get_client") as get_client_mock,
            patch("apps.core.ai_assistant._call_llm") as call_llm_mock,
            patch("apps.core.ai_assistant.connection", mock_connection),
        ):
            result = ai_assistant.ask_ai(self.PIE_CHART_PROMPT, limit=10)

        self.assertTrue(result["ok"])
        self.assertTrue(result["chart_requested"])
        self.assertIsNotNone(result["chart"])
        self.assertEqual(result["chart"]["type"], "pie")
        self.assertEqual(result["chart"]["labels"], ["Gen", "Mar"])
        self.assertEqual(result["chart"]["datasets"][0]["data"], [2, 1])
        self.assertTrue(result["chart"]["showPercentages"])
        self.assertEqual(result["chart"]["datasets"][0]["percentages"], [67, 33])
        call_llm_mock.assert_not_called()
        get_client_mock.assert_not_called()

    def test_ask_ai_returns_line_chart_for_andamento_prompt(self):
        mock_cursor = MagicMock()
        mock_cursor.description = [
            SimpleNamespace(name="ID"),
            SimpleNamespace(name="DataReg"),
            SimpleNamespace(name="Avere_Imponibile"),
        ]
        mock_cursor.fetchmany.return_value = [(1, "2026-01-15", 16000)]
        mock_cursor.fetchall.side_effect = [
            [(1, 2), (3, 1)],
            [(1, 2), (3, 1)],
        ]
        mock_connection = MagicMock()
        mock_connection.cursor.return_value.__enter__.return_value = mock_cursor
        with (
            patch("apps.core.ai_assistant._get_client") as get_client_mock,
            patch("apps.core.ai_assistant._call_llm") as call_llm_mock,
            patch("apps.core.ai_assistant.connection", mock_connection),
        ):
            result = ai_assistant.ask_ai(self.LINE_CHART_PROMPT, limit=10)

        self.assertTrue(result["ok"])
        self.assertIsNotNone(result["chart"])
        self.assertEqual(result["chart"]["type"], "line")
        call_llm_mock.assert_not_called()
        get_client_mock.assert_not_called()
